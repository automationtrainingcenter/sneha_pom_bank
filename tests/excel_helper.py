from openpyxl import load_workbook

workbook_name = ''
sheet_name = ''


# open a workbook to read data
def open_workbook(workbook, sheet):
    global workbook_name, sheet_name
    workbook_name = workbook
    sheet_name = sheet
    wb = load_workbook(filename=workbook_name)


# count the no of rows
def get_row_count():
    global workbook_name, sheet_name
    file_path = workbook_name
    wb = load_workbook(file_path)
    return wb[sheet_name].max_row


def get_column_count():
    global workbook_name, sheet_name
    file_path = workbook_name
    wb = load_workbook(file_path)
    return wb[sheet_name].max_column


def read_cell_data(row_num, col_num):
    global workbook_name, sheet_name
    file_path = workbook_name
    wb = load_workbook(file_path)
    return wb[sheet_name].cell(row_num, col_num).value


# open_workbook('E:\SMIT\sneha\sneha_bank\TestData.xlsx', 'BranchData')
# for row in range(2, get_row_count() + 1):
#     for col in range(1, get_column_count() + 1):
#         cell = read_cell_data(row, col)
#         print(cell, end='\t')
#     print()
